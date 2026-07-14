"""Query logs API router."""

import html
import io
import re
from datetime import datetime, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, select, desc

from src.logs.models import QueryLog
from admin.dependencies import get_db
from admin.schemas.query_logs import QueryLogList, QueryLogItem

router = APIRouter(prefix="/journal", tags=["Journal"])

# ── Export tunables / labels ────────────────────────────────────────────────
_MAX_EXPORT_ROWS = 50_000          # hard cap so a full-history export can't OOM
_MAX_CELL = 32_000                 # Excel's per-cell limit is 32767
_TAG_RE = re.compile(r"<[^>]+>")
# XML 1.0 forbids these control chars; openpyxl raises on them — strip first.
_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

_QTYPE_LABEL = {
    "RAG_PRODUCT": "RAG (товар)",
    "FAQ_EXACT": "FAQ",
    "FAQ_COMPANY": "О компании",
    "FAQ_DEALER": "Дилер",
    "ERROR": "Ошибка",
}
_FEEDBACK_LABEL = {
    "good": "👍 Полезно",
    "bad": "👎 Не помогло",
    "operator": "🆘 Оператор",
}


def _apply_filters(q, query_type: Optional[str], feedback: Optional[str], search: Optional[str]):
    """Shared filter clause for both the list view and the export."""
    if query_type:
        q = q.where(QueryLog.query_type == query_type)
    if feedback == "none":
        q = q.where(QueryLog.feedback.is_(None))
    elif feedback:
        q = q.where(QueryLog.feedback == feedback)
    if search:
        q = q.where(func.lower(QueryLog.question).contains(search.lower()))
    return q


def _clean(v) -> str:
    """Any DB value → an Excel-safe string (illegal control chars removed, capped)."""
    if v is None:
        return ""
    return _ILLEGAL_XML_RE.sub("", str(v))[:_MAX_CELL]


def _to_plain(s: Optional[str]) -> str:
    """HTML-ish answer → readable plain text for a spreadsheet cell."""
    if not s:
        return ""
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.IGNORECASE)
    s = _TAG_RE.sub("", s)
    s = html.unescape(s)
    return _ILLEGAL_XML_RE.sub("", s).strip()[:_MAX_CELL]


@router.get("", response_model=QueryLogList)
async def list_query_logs(
    query_type: Optional[str] = Query(None),
    feedback: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """List query logs with filtering and pagination."""
    q = _apply_filters(select(QueryLog), query_type, feedback, search)

    total = db.execute(select(func.count()).select_from(q.subquery())).scalar()

    offset = (page - 1) * per_page
    rows = db.execute(
        q.order_by(desc(QueryLog.ts)).offset(offset).limit(per_page)
    ).scalars().all()

    return QueryLogList(
        items=[QueryLogItem.model_validate(r) for r in rows],
        total=total,
        page=page,
        per_page=per_page,
    )


@router.get("/export")
async def export_query_logs(
    query_type: Optional[str] = Query(None),
    feedback: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="ГГГГ-ММ-ДД, в часовом поясе клиента (включительно)"),
    date_to: Optional[str] = Query(None, description="ГГГГ-ММ-ДД, включительно"),
    tz_offset_min: int = Query(0, description="минуты, прибавляемые к UTC для получения локального времени клиента"),
    db: Session = Depends(get_db),
):
    """Выгрузить журнал (вопрос/ответ + 👍/👎 + замечание + судья) в Excel (.xlsx).

    Даты трактуются в часовом поясе клиента (`tz_offset_min`), а `ts` в БД — UTC:
    границы дня переводятся в UTC для фильтра, а время в отчёте — обратно в локальное.
    """
    # Lazy import so a missing openpyxl only breaks this endpoint, not admin boot.
    from openpyxl import Workbook
    from openpyxl.cell.cell import WriteOnlyCell
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    q = _apply_filters(select(QueryLog), query_type, feedback, search)

    tz = timedelta(minutes=tz_offset_min)

    def _parse(d: str) -> datetime:
        try:
            return datetime.strptime(d.strip(), "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=422, detail=f"Неверная дата {d!r} — нужен формат ГГГГ-ММ-ДД")

    if date_from:
        q = q.where(QueryLog.ts >= _parse(date_from) - tz)
    if date_to:
        q = q.where(QueryLog.ts < (_parse(date_to) + timedelta(days=1)) - tz)

    rows = db.execute(
        q.order_by(QueryLog.ts.asc()).limit(_MAX_EXPORT_ROWS)
    ).scalars().all()

    wb = Workbook(write_only=True)          # streaming writer — memory-safe for big exports
    ws = wb.create_sheet("Журнал")

    headers = [
        "ID", "Дата и время", "Пользователь", "Тип", "Вопрос", "Переформулировка",
        "Ответ", "Оценка", "Замечание", "Судья (0-100)", "Вердикт судьи",
        "Score", "Чанки", "Город",
    ]
    widths = [7, 18, 16, 13, 46, 40, 64, 14, 40, 12, 44, 8, 7, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    bold = Font(bold=True)
    header_cells = []
    for h in headers:
        c = WriteOnlyCell(ws, value=h)
        c.font = bold
        header_cells.append(c)
    ws.append(header_cells)

    for r in rows:
        local_ts = (r.ts + tz) if r.ts else None
        user = ("@" + r.username) if r.username else (r.user_id if r.user_id is not None else "")
        ws.append([
            r.id,
            local_ts.strftime("%Y-%m-%d %H:%M:%S") if local_ts else "",
            _clean(user),
            _QTYPE_LABEL.get(r.query_type, r.query_type or ""),
            _clean(r.question),
            _clean(r.reformulated_query),
            _to_plain(r.answer),
            _FEEDBACK_LABEL.get(r.feedback, ""),
            _clean(r.feedback_note),
            r.usefulness_score if r.usefulness_score is not None else "",
            _clean(r.usefulness_verdict),
            round(r.top_score, 3) if r.top_score is not None else "",
            r.chunks_used if r.chunks_used is not None else "",
            _clean(r.city),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # ASCII-only filename — HTTP headers are latin-1, Cyrillic would 500.
    span = f"{date_from or 'all'}_{date_to or 'all'}"
    filename = f"teplodar-journal_{span}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{log_id}/context", response_model=List[QueryLogItem])
async def get_dialog_context(
    log_id: int,
    limit: int = Query(3, ge=1, le=10),
    window_minutes: int = Query(30, ge=1, le=240),
    db: Session = Depends(get_db),
):
    """Return up to `limit` previous Q&A turns from the same user_id
    within `window_minutes` before this log entry. Used by the Journal
    detail panel to show what the user asked before this turn.
    """
    target = db.get(QueryLog, log_id)
    if target is None:
        raise HTTPException(status_code=404, detail="Log entry not found")
    if not target.user_id:
        return []

    cutoff = target.ts - timedelta(minutes=window_minutes)
    prev = db.execute(
        select(QueryLog)
        .where(QueryLog.user_id == target.user_id)
        .where(QueryLog.id != target.id)
        .where(QueryLog.ts < target.ts)
        .where(QueryLog.ts >= cutoff)
        .order_by(desc(QueryLog.ts))
        .limit(limit)
    ).scalars().all()

    # Return oldest first so the UI reads top-down like a chat thread
    return [QueryLogItem.model_validate(r) for r in reversed(prev)]
